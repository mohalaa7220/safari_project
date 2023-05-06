from rest_framework import generics, status, views
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .serializers import ReportSerializer, AddReportSerializer, UpdateReportSerializer
from .models import Report
import io
import openpyxl
import os
from django.conf import settings
from django.urls import reverse
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from project.serializer_error import serializer_error
from django.http import HttpResponseRedirect


class AddReport(generics.ListCreateAPIView):
    queryset = Report.objects.select_related('added_by', 'product').all()
    serializer_class = ReportSerializer
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = AddReportSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(added_by=request.user)
            return Response({'message': "Report Created successfully"}, status=status.HTTP_200_OK)
        else:
            return serializer_error(serializer)


class DetailsReport(generics.RetrieveUpdateDestroyAPIView):
    queryset = Report.objects.select_related('added_by', 'product').all()
    serializer_class = ReportSerializer
    permission_classes = [IsAuthenticated]

    def update(self, request, pk=None):
        report = self.get_object()
        serializer = UpdateReportSerializer(instance=report, data=request.data)
        if serializer.is_valid():
            serializer.save(added_by=request.user)
            return Response({'message': "Report Updated successfully"}, status=status.HTTP_200_OK)
        else:
            return serializer_error(serializer)


def report_download(request, report_id=None):
    if report_id is not None:
        # Download a single report
        report = get_object_or_404(Report, id=report_id)
        filename = "report_" + \
            str(report.id) + "_" + \
            report.product.name + ".xlsx"

        # Generate Excel file in memory
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Report"

        worksheet.cell(row=1, column=1, value="ID")
        worksheet.cell(row=1, column=2, value="Text")
        worksheet.cell(row=1, column=3, value="Product Name")
        worksheet.cell(row=1, column=4, value="Price")
        worksheet.cell(row=1, column=5, value="Quantity")
        worksheet.cell(row=1, column=6, value="In Stock")
        worksheet.cell(row=1, column=7, value="Created")

        worksheet.cell(row=2, column=1, value=str(report.id))
        worksheet.cell(row=2, column=2, value=report.text)
        worksheet.cell(row=2, column=3, value=report.product.name)
        worksheet.cell(row=2, column=4, value=report.product.price)
        worksheet.cell(row=2, column=5, value=report.product.quantity)
        worksheet.cell(row=2, column=6, value=report.product.in_stock)
        worksheet.cell(row=2, column=7,
                       value=report.created.strftime("%Y-%m-%d %H:%M:%S"))

        column_widths = [15, 50, 30, 15, 15, 20]
        for i, width in enumerate(column_widths):
            column = worksheet.column_dimensions[openpyxl.utils.get_column_letter(
                i+1)]
            column.width = width

        # Save workbook to memory
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Serve file to client
        response = HttpResponse(
            buffer, content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    else:
        # Download all reports
        filename = "all_reports.xlsx"
        filepath = os.path.join(settings.MEDIA_ROOT, filename)

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Reports"

        worksheet.cell(row=1, column=1, value="ID")
        worksheet.cell(row=1, column=2, value="Text")
        worksheet.cell(row=1, column=3, value="Product Name")
        worksheet.cell(row=1, column=4, value="Price")
        worksheet.cell(row=1, column=5, value="Quantity")
        worksheet.cell(row=1, column=6, value="In Stock")
        worksheet.cell(row=1, column=7, value="Created")

        row = 2
        for report in Report.objects.select_related('added_by', 'product').all():
            worksheet.cell(row=row, column=1, value=str(report.id))
            worksheet.cell(row=row, column=2, value=report.text)
            worksheet.cell(row=row, column=3, value=report.product.name)
            worksheet.cell(row=row, column=4, value=report.product.price)
            worksheet.cell(row=row, column=5, value=report.product.quantity)
            worksheet.cell(row=row, column=6, value=report.product.in_stock)
            worksheet.cell(row=row, column=7,
                           value=report.created.strftime("%Y-%m-%d %H:%M:%S"))
            row += 1
        # Apply styles to columns
        column_widths = [15, 50, 30, 15, 15, 20]
        for i, width in enumerate(column_widths):
            column = worksheet.column_dimensions[openpyxl.utils.get_column_letter(
                i+1)]
            column.width = width

        workbook.save(filepath)

        # Return response with Excel file as a download
        with open(filepath, 'rb') as f:
            response = HttpResponse(
                f.read(), content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


class ExportExcel(views.APIView):

    def get(self, request, pk=None):
        report = Report.objects.select_related('user', 'product').get(pk=pk)
        filename = "report_" + \
            str(report.id) + "_" + \
            report.product.name + ".xlsx"
        filepath = os.path.join(settings.MEDIA_ROOT, filename)

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Report"

        worksheet.cell(row=1, column=1, value="ID")
        worksheet.cell(row=1, column=2, value="Text")
        worksheet.cell(row=1, column=3, value="Product Name")
        worksheet.cell(row=1, column=4, value="Price")
        worksheet.cell(row=1, column=5, value="Quantity")
        worksheet.cell(row=1, column=6, value="In Stock")
        worksheet.cell(row=1, column=7, value="Created")

        worksheet.cell(row=1, column=1, value=str(report.id))
        worksheet.cell(row=1, column=2, value=report.text)
        worksheet.cell(row=1, column=3, value=report.product.name)
        worksheet.cell(row=1, column=4, value=report.product.price)
        worksheet.cell(row=1, column=5, value=report.product.quantity)
        worksheet.cell(row=1, column=6, value=report.product.in_stock)
        worksheet.cell(row=1, column=7,
                       value=report.created.strftime("%Y-%m-%d %H:%M:%S"))

        workbook.save(filepath)

        # Generate download link
        download_link = request.build_absolute_uri(
            reverse('report_download', args=[report.id]))

        # Return response with Excel file as a download
        response = HttpResponseRedirect(download_link)

        return response
